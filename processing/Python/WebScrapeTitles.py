import time
from BeautifulSoup import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook

file = r"C:\temp\Google_UpdateOnly_623.xlsx"
wb = load_workbook(file)
ws = wb.active

rwCnt = ws.get_highest_row()

browser = webdriver.Chrome(r"C:\Share\chromedriver.exe")

for row in ws.iter_rows('A2:A{}'.format(rwCnt)):
    cell = row[0]
    url = cell.value
    browser.get(url)
    time.sleep(3)
    data = browser.page_source
    soup = BeautifulSoup(data)
    title = soup.find('title')
    ws['B{}'.format(cell.row)] = title.renderContents()

wb.save(file)
browser.quit()

